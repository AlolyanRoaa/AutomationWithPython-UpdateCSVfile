import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def proccess_workBook(filename):
    workBook = xl.load_workbook(filename) #load the excel file
    sheet = workBook["Sheet1"] #to access the sheet spcify the nam eof that sheet
    cell = sheet["a1"] #access the cell by its cordinate == sheet.cell(1,1)
    print(cell.value)
    print(sheet.max_row) #number of rows in that sheet
    print("\n")

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row , 3) # to access the price column
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4) #create a new cell ro store the new price
        corrected_price_cell.value = corrected_price # store the new value to the cell

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    workBook.save(filename)



